from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Literal, Optional, cast, Mapping

import pandas as pd
import yaml


SchemaName = Literal["old", "new"]


from ..types import SchemaConfig

_schema_cache: SchemaConfig | None = None


def _load_schema() -> SchemaConfig:
    """Load simplified schema.yaml."""
    global _schema_cache
    if _schema_cache is None:
        schema_path = Path("configs/schema.yaml")
        with open(schema_path, "r", encoding="utf-8") as f:
            _schema_cache = cast(SchemaConfig, yaml.safe_load(f))
    return _schema_cache


@dataclass(frozen=True)
class InputData:
    df: pd.DataFrame
    schema: SchemaName


@dataclass(frozen=True)
class InputSheet:
    name: str
    df: pd.DataFrame
    schema: SchemaName
    header_row_excel: int  # 1-based
    first_data_index: int  # pandas index of first data row


@dataclass(frozen=True)
class SkippedSheet:
    name: str
    reason: str


def read_results_excel(path: Path) -> InputData:
    df = pd.read_excel(path)
    schema = detect_schema(df)
    return InputData(df=df, schema=schema)


def detect_schema(df: pd.DataFrame) -> SchemaName:
    """Detect schema with stricter heuristics.

    - new: requires Species/Compound/Class/MatchScore
    - old: accepts long Comments header OR a minimal set of expected columns
    - otherwise: raise error (so caller can skip the sheet)
    """
    cols_lower = {str(c).lower() for c in df.columns}
    # New schema: has these 4 key columns
    new_markers = {"species", "compound", "class", "matchscore"}
    if all(marker in cols_lower for marker in new_markers):
        return "new"
    # Old schema: long Comments header or minimal old columns present
    long_comments = "comments (note here, for example, if there are common names and official iupac names that are actually the same compound)"
    if any(long_comments in str(c).lower() for c in df.columns):
        return "old"
    minimal_old = {"daterun", "cartridgenum", "match1"}
    if minimal_old.issubset(cols_lower):
        return "old"
    raise ValueError("no_schema_detected")


def write_excel(sheets: Mapping[str, pd.DataFrame], out_path: Path) -> None:
    """Write multiple DataFrames to Excel as native Tables with default Excel style.

    - Ensures new-schema column order
    - Adds an Excel Table over the data range
    - Uses a standard built-in style close to Excel's default (Medium 2)
    """
    schema = _load_schema()
    order: list[str] = schema["new_schema"]

    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            # Ensure all columns exist (add empty if missing)
            for col in order:
                if col not in df.columns:
                    df[col] = pd.NA
            df_out = df[order]

            # Plain write
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)

            # Wrap the written range in an Excel Table with a default built-in style
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter

            ws = writer.sheets[sheet_name]
            max_row = len(df_out) + 1  # +1 for header
            max_col = len(df_out.columns)
            col_letter = get_column_letter(max_col)
            table_ref = f"A1:{col_letter}{max_row}"

            table = Table(displayName=sheet_name.replace(" ", "_"), ref=table_ref)
            # Choose a common default style that matches Excel's default look closely
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

            # Make columns slightly wider for readability (no fancy auto-fit)
            headers = list(df_out.columns)
            for idx, header in enumerate(headers, start=1):
                col = get_column_letter(idx)
                # Base width on header length with a sensible minimum and cap
                base = len(str(header)) + 2
                width = max(18, min(40, base))
                ws.column_dimensions[col].width = width

                # Header styling: white text, left-aligned
                cell = ws.cell(row=1, column=idx)
                cell.alignment = Alignment(horizontal="left")
                cell.font = Font(color="FFFFFFFF")


def read_mapping_excel(path: Optional[Path]) -> Optional[pd.DataFrame]:
    if path is None:
        return None
    return pd.read_excel(path)


def _row_matches_headers(row_vals: list[str], required: list[str]) -> bool:
    """Check if row contains all required headers (case-insensitive),
    allowing configured alias variants (e.g., long Comments header).

    This helps detect tables where a header uses an alias rather than
    the canonical name defined in the schema.
    """
    row_set = {v.strip().lower() for v in row_vals if isinstance(v, str)}

    # Expand row_set with canonical names for any alias variants present
    try:
        schema = _load_schema()
        aliases = schema.get("aliases", {}) or {}
        for canonical, variants in aliases.items():
            for variant in variants:
                if isinstance(variant, str) and variant.strip().lower() in row_set:
                    row_set.add(str(canonical).strip().lower())
    except Exception:
        # Best-effort: if schema/aliases unavailable, skip alias expansion
        pass

    return all(col.lower() in row_set for col in required)


_SCAN_LIMIT = 200


def _find_table(df_raw: pd.DataFrame) -> pd.DataFrame | None:
    """Find table by scanning the first _SCAN_LIMIT rows for header row."""
    schema = _load_schema()
    new_schema: list[str] = schema["new_schema"]
    old_schema: list[str] = schema["old_schema"]

    for idx in range(min(len(df_raw), _SCAN_LIMIT)):
        row = [str(v) if not isinstance(v, str) else v for v in df_raw.iloc[idx].tolist()]
        if _row_matches_headers(row, new_schema):
            df = df_raw.iloc[idx + 1 :].copy()
            df.columns = row
            return df
        if _row_matches_headers(row, old_schema):
            df = df_raw.iloc[idx + 1 :].copy()
            df.columns = row
            return df
    return None


def read_results_workbook_detailed(path: Path) -> tuple[list[InputSheet], list[SkippedSheet]]:
    xls = pd.ExcelFile(path)
    sheets: list[InputSheet] = []
    skipped: list[SkippedSheet] = []
    for name in xls.sheet_names:
        try:
            raw = pd.read_excel(xls, sheet_name=name, header=None, dtype=object)
            table = _find_table(raw)
            if table is None:
                skipped.append(SkippedSheet(name=name, reason="no_table_header"))
                continue
            else:
                try:
                    schema = detect_schema(table)
                    # compute header index by scanning again for robustness
                    header_idx = None
                    schema_conf = _load_schema()
                    for idx in range(min(len(raw), _SCAN_LIMIT)):
                        row = [
                            str(v) if not isinstance(v, str) else v for v in raw.iloc[idx].tolist()
                        ]
                        if _row_matches_headers(
                            row, schema_conf["new_schema"]
                        ) or _row_matches_headers(row, schema_conf["old_schema"]):
                            header_idx = idx
                            break
                    if header_idx is None:
                        skipped.append(SkippedSheet(name=name, reason="header_row_not_found"))
                        continue
                    sheets.append(
                        InputSheet(
                            name=name,
                            df=table,
                            schema=schema,
                            header_row_excel=header_idx + 1,
                            first_data_index=header_idx + 1,
                        )
                    )
                except Exception as e:
                    skipped.append(SkippedSheet(name=name, reason=f"no_schema: {e}"))
        except Exception as e:
            skipped.append(SkippedSheet(name=name, reason=f"read_error: {e}"))
            continue
    return sheets, skipped


def read_results_workbook(path: Path) -> list[InputSheet]:
    sheets, _ = read_results_workbook_detailed(path)
    return sheets


def read_mapping_workbook(path: Optional[Path]) -> Optional[pd.DataFrame]:
    if path is None:
        return None
    xls = pd.ExcelFile(path)
    frames: list[pd.DataFrame] = []
    for name in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=name)
            cols = {c.lower(): c for c in df.columns}
            if {"daterun", "cartridgenum", "species"}.issubset(cols.keys()):
                frames.append(
                    df[[cols["daterun"], cols["cartridgenum"], cols["species"]]].rename(
                        columns={
                            cols["daterun"]: "DateRun",
                            cols["cartridgenum"]: "CartridgeNum",
                            cols["species"]: "Species",
                        }
                    )
                )
        except Exception:
            continue
    if frames:
        return pd.concat(frames, ignore_index=True)
    return None


class IOService:
    def __init__(self, logger: logging.Logger) -> None:
        self.logger = logger

    def read_results(self, path: Path) -> InputData:
        self.logger.info("Reading results", extra={"path": str(path)})
        inp = read_results_excel(path)
        self.logger.info(
            "Parsed results (single sheet)", extra={"schema": inp.schema, "rows": len(inp.df)}
        )
        return inp

    def read_results_multi(self, path: Path) -> list[InputSheet]:
        self.logger.info("Reading workbook (multi-sheet)", extra={"path": str(path)})
        sheets, skipped = read_results_workbook_detailed(path)
        self.logger.info(
            "Parsed workbook",
            extra={
                "sheet_count": len(sheets),
                "sheets": [s.name for s in sheets],
                "schemas": {s.name: s.schema for s in sheets},
            },
        )
        for sk in skipped:
            self.logger.info("Skipped sheet", extra={"sheet": sk.name, "reason": sk.reason})
        return sheets

    def read_results_multi_detailed(
        self, path: Path
    ) -> tuple[list[InputSheet], list[SkippedSheet]]:
        self.logger.info("Reading workbook (multi-sheet)", extra={"path": str(path)})
        sheets, skipped = read_results_workbook_detailed(path)
        return sheets, skipped

    def read_mapping(self, path: Optional[Path]) -> Optional[pd.DataFrame]:
        if path is None:
            self.logger.info("No mapping provided")
            return None
        self.logger.info("Reading mapping", extra={"path": str(path)})
        df = read_mapping_workbook(path)
        self.logger.info("Parsed mapping", extra={"rows": 0 if df is None else len(df)})
        return df

    def write_output(self, sheets: Mapping[str, pd.DataFrame], out_path: Path) -> None:
        self.logger.info(
            f"Writing {out_path.name}", extra={"path": str(out_path), "sheets": len(sheets)}
        )
        write_excel(sheets, out_path)
