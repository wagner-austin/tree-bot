from __future__ import annotations

from pathlib import Path
from typing import List

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..aggregate.summary import Section


def write_summary_into_workbook(std_path: Path, sections: List[Section]) -> None:
    """Append/replace a 'Summary' sheet in an existing standardized workbook.

    Each section is labeled 'Site: <site> | Species: <species> (unique_compounds=..., total_peaks=...)'
    followed by a table with ordered columns: Compound, Compound Class, RetentionMin, RetentionMax,
    RtRange, Count, Comments.
    """
    with pd.ExcelWriter(std_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        sheet_name = "Summary"
        # Start with an empty frame so the sheet exists
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        current_row = 1

        def _write_header(text: str) -> None:
            nonlocal current_row
            cell = ws.cell(row=current_row, column=1, value=text)
            cell.font = Font(bold=True)
            current_row += 1

        for section in sections:
            # Section header with clearer labels: kept metrics
            hdr = (
                f"Site: {section.site} | Species: {section.species} "
                f"(unique_compounds_kept={section.stats.get('unique_compounds', 0)}, "
                f"total_compounds={section.stats.get('unique_compounds_all', 0)})"
            )
            _write_header(hdr)

            # Write table starting at current_row
            start_row = current_row
            # Ensure requested column order
            cols = [
                "Compound",
                "Compound Class",
                "RetentionMin",
                "RetentionMax",
                "RtRange",
                "AvgMatchQuality",
                "Count",
                "Comments",
            ]
            df = section.df.reindex(columns=cols)

            # Write headers manually
            for col_idx, header in enumerate(cols, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFFFF")
                cell.alignment = Alignment(horizontal="left")

            # Write data rows manually
            for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
                for col_idx, col_name in enumerate(cols, start=1):
                    value = row[col_name]
                    # Handle NaN/None values
                    if pd.isna(value):
                        value = None
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Wrap as Excel Table
            max_row = start_row + len(df)
            max_col = len(df.columns)
            ref = f"A{start_row}:{get_column_letter(max_col)}{max_row}"
            display = f"{section.site}_{section.species}".replace(" ", "_")[:31]
            table = Table(displayName=display, ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

            # Basic column width
            for idx, header in enumerate(cols, start=1):
                col = get_column_letter(idx)
                base = max(len(str(header)), 12) + 2
                ws.column_dimensions[col].width = min(40, base)

            # One blank line gap
            current_row = max_row + 2
